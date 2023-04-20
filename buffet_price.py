from fpdf import FPDF
import openpyxl


class BuffetPricesGenerator:
    def __init__(self):
        self.pageW = 918
        self.pageH = 691

        self.DRINKS_COLOR_RGB = (103, 59, 95)
        self.DESSERTS_COLOR_RGB = (240, 93, 112)
        self.BREAKFAST_COLOR_RGB = (111, 153, 149)
        self.HOTTER_COLOR_RGB = (232, 99, 82)
        self.SALADS_COLOR_RGB = (152, 185, 106)
        self.BAKED_GOODS = (89, 91, 129)
        self.active_bg_color = self.DRINKS_COLOR_RGB

        self.pdf = FPDF(format=(self.pageW, self.pageH), unit="mm")
        self.pdf.add_font("BebasNeue", "", "./BebasNeueBold.ttf", uni=True)

        self.exel_file = openpyxl.load_workbook('prices.xlsx')

    def draw_background(self):
        self.pdf.set_line_width(1)
        self.pdf.set_fill_color(self.active_bg_color[0], self.active_bg_color[1], self.active_bg_color[2])
        self.pdf.rect(0, 0, self.pageW, self.pageH, 'F')

    def draw_title(self, txt: str):
        self.pdf.set_font("BebasNeue", size=290)
        self.pdf.set_text_color(255, 255, 255)
        self.pdf.set_xy(67, 50)
        self.pdf.multi_cell(self.pageW - 200, 115, txt, 0, "L")

    def draw_circle(self):
        self.pdf.set_fill_color(255, 255, 255)
        self.pdf.ellipse(540, 299, 413, 413, 'F')

    def draw_logo(self):
        self.pdf.image("./bg/buffet_logo.png", h=50, w=201, x=70, y=560)

    def draw_price(self, price: str):
        self.pdf.set_font("BebasNeue", size=451)
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.set_xy(129, 427)
        self.pdf.multi_cell(self.pageW - 200, 115, price, 0, "R")

    def create_page(self, title: str, price: str):
        self.pdf.add_page()
        self.draw_background()
        self.draw_title(title)
        self.draw_circle()
        self.draw_price(price)
        self.draw_logo()

    def change_active_bg_color(self, price_group: str):
        price_group = price_group.lower()

        if price_group == "горячие блюда:":
            self.active_bg_color = self.HOTTER_COLOR_RGB
        elif price_group == "затраки:":
            self.active_bg_color = self.BREAKFAST_COLOR_RGB
        elif price_group == "напитки:":
            self.active_bg_color = self.DRINKS_COLOR_RGB
        elif price_group == "десерты:":
            self.active_bg_color = self.DESSERTS_COLOR_RGB
        elif price_group == "салаты :":
            self.active_bg_color = self.SALADS_COLOR_RGB
        elif price_group == "выпечка :":
            self.active_bg_color = self.BAKED_GOODS

    def generate_prices_from_exel_at_pdf(self):
        worksheet = self.exel_file.active
        for col in worksheet.iter_rows(min_col=1, max_col=2, max_row=worksheet.max_row, min_row=3, values_only=True):
            if col[0] is not None and col[1] is not None:
                self.create_page(str(col[0]), str(col[1]))
            elif col[0] is not None and col[1] is None:
                self.change_active_bg_color(col[0])

        self.pdf.output("simple_demo.pdf")


if __name__ == '__main__':
    generator = BuffetPricesGenerator()
    generator.generate_prices_from_exel_at_pdf()
